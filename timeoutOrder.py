# -- coding: utf-8 --
# @Time : 2024/6/14 18:09
# @Author : Eagle Yin
import requests
import json
import xlrd
from datetime import datetime,timedelta
from openpyxl import Workbook

# 查询任务
taskQuery_url = 'http://10.79.48.14:9003/adapter-api/monitor/task/query'
# 查询所有位置的接口
queryLocation_url = 'http://192.168.29.128:9000/model/queryModelByType?modelType=location'
# 请求头
headers = {'Content-Type': 'application/json', 'accept': 'application/json'}

# 查询任务
def getTask(b):
    task = []
    for i in range(len(b)):
        d = b[i]
        r = json.dumps(d)
        response = requests.post(taskQuery_url, headers=headers, data=r)
        # print(response.json())
        tasks = response.json().get('data').get('tasks')
        for a in tasks:
            taskCode = a.get('taskCode')
            status = a.get('status')
            containerCode = a.get('containerCode')
            toLocationCode = a.get('toLocationCode')
            deadline = a.get('deadline')
            taskType = a.get('taskType')
            zone = a.get('zone')
            taskPriority = a.get('taskPriority')
            # 创建任务对象（也可以使用字典）
            new_task = {'taskCode': taskCode, 'status': status,'containerCode':containerCode,'toLocationCode':toLocationCode,'deadline':deadline,'taskType':taskType,'zone':zone,"taskPriority":taskPriority}
            # 将任务添加到新的任务列表中
            task.append(new_task)
    return task

# 查询接近超时任务
def getoutTimeTask(c):
    # 创建一个新的任务列表来存储截止时间小于30分钟的任务
    tasks_less_than_30_min = []
    # 任务类型英文到中文的映射
    taskType_mapping = {
        "TMS": "出库",
        "PROCESSING":"执行中",
        "PENDING":"待办"
    }
    # 获取当前时间
    now = datetime.now()

    # 遍历任务列表
    for task in c:
        # 提取截止时间
        deadline = task["deadline"]
        deadline_time = datetime.strptime(deadline, '%Y-%m-%d %H:%M:%S')
        # 检查截止时间是否小于当前时间加上30分钟
        if (deadline_time - now).total_seconds() / 60 < 30 and deadline_time > now:
            #转换任务状态和任务类型
            task["taskType"] = taskType_mapping[task["taskType"]]
            task["status"] = taskType_mapping[task["status"]]
            # 如果满足条件，则添加到新列表中
            tasks_less_than_30_min.append(task)
            # 将新列表转换为JSON格式的字符串
    # json_tasks_less_than_30_min = json.dumps(tasks_less_than_30_min, ensure_ascii=False, indent=4)

    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks Less Than 30 Minutes"

    # 设置表头
    headers = ["任务号", "状态", "容器号","工作站","截单时间","任务类型","子仓","优先级"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

        # 将提取的任务写入Excel工作表
    for row_num, task in enumerate(tasks_less_than_30_min, 2):
        for col_num, (key, value) in enumerate(task.items(), 1):
            ws.cell(row=row_num, column=col_num, value=value)

            # 保存Excel文件
    wb.save("tasks_less_than_30_minutes.xlsx")


# 翻页列表
def page(i):
    pageList = []
    for page in range(i):
        b = {"page":page+1,"size":50,"statuses":["PROCESSING","PENDING"],"taskType":"TMS"}
        pageList.append(b)
    return pageList


if __name__ == '__main__':
    b = page(100)
    # print(b)
    c = getTask(b)
    getoutTimeTask(c)
