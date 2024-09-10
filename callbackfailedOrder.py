# -- coding: utf-8 --
# @Time : 2024/6/14 18:09
# @Author : Eagle Yin
import requests
import json
from openpyxl import Workbook

# 查询任务
taskQuery_url = 'http://10.79.48.14:9003/ess-api/callback/query?pageIndex=1&pageSize=20'
# 查询所有位置的接口
queryLocation_url = 'http://192.168.29.128:9000/model/queryModelByType?modelType=location'
# 请求头
headers = {'Content-Type': 'application/json', 'accept': 'application/json'}

# 查询任务
def query_task(b):
    task = []
    for i in range(len(b)):
        d = b[i]
        r = json.dumps(d)
        response = requests.post(taskQuery_url, headers=headers, data=r)
        # 提取callbacks列表
        callbacks = response.json()['data']['callbacks']
        # 遍历callbacks列表，并提取taskCode、state和containerCode
        for callback in callbacks:
            # 解析message字段中的JSON字符串
            message_data = json.loads(callback['message'])

            # 提取所需字段
            taskCode = message_data['taskCode']
            state = callback['state']
            containerCode = message_data['containerCode']

            # 创建任务对象（也可以使用字典）
            new_task = {'taskCode': taskCode, 'containerCode': containerCode,'state':state}
            task.append(new_task)

            # 打印或处理这些字段
            print(f"taskCode: {taskCode}, state: {state}, containerCode: {containerCode}")
    return task


# 存储回传失败的任务到excel
def getcallbackTask(c):
    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "TasksFailed"

    # 设置表头
    headers = ["任务号", "状态", "容器号"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

        # 将提取的任务写入Excel工作表
    for row_num, task in enumerate(c, 2):
        for col_num, (key, value) in enumerate(task.items(), 1):
            ws.cell(row=row_num, column=col_num, value=value)

            # 保存Excel文件
    wb.save("TasksFailed.xlsx")


# 翻页列表
def page():
    pageList = []
    # 获取总条数
    Payload = json.dumps({"isAbnormal": False, "name": "CALLBACK_OF_TOTE_LOADED_BY_ROBOT", "taskCode": "TGXJ24091001", "limit": 20,
     "page": 1})
    response = requests.post(taskQuery_url, headers=headers, data=Payload)
    total = response.json().get('data')['total']
    # 计算总页数
    total_pages = (total - 1) // 20 + 1
    for page in range(total_pages):
        Payload_new = {"isAbnormal":False,"name":"CALLBACK_OF_TOTE_LOADED_BY_ROBOT","taskCode":"TGXJ24091001","limit":20,"page":page+1}
        pageList.append(Payload_new)
    return pageList


if __name__ == '__main__':
    b = page()
    c = query_task(b)
    getcallbackTask(c)

